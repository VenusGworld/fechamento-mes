import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment

from config import caminhos_entrada, caminhos_saida, classes


def interpretar_planilha(tipo):
    caminho = caminhos_entrada.get(tipo)
    classe = classes.get(tipo)

    with open(caminho, 'r', encoding='UTF8') as arquivo:
        leitor_csv = csv.reader(arquivo)
        next(leitor_csv)
        registros = [classe(linha) for linha in leitor_csv]

    return registros


def inicializar_arquivo(tipo, titulo):
    caminho = caminhos_saida.get(tipo)

    workbook = Workbook()
    sheet = workbook.active

    sheet.merge_cells('A1:D1')
    sheet['A1'] = titulo
    sheet['E1'] = 'Justificativa'

    workbook.save(caminho)


def gravar_linha(tipo, *args):
    caminho = caminhos_saida.get(tipo)

    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook.active

    proxima_linha = len(sheet['A']) + 1

    if len(args) == 1:
        celula_inicial = sheet.cell(row=proxima_linha, column=1, value=args[0])
        celula_final = sheet.cell(row=proxima_linha, column=4)
        sheet.merge_cells(f'{celula_inicial.coordinate}:{celula_final.coordinate}')

    if len(args) == 2:
        celula_inicial = sheet.cell(row=proxima_linha, column=1, value=args[0])
        celula_final = sheet.cell(row=proxima_linha, column=3)
        sheet.cell(row=proxima_linha, column=4, value=args[1])
        sheet.merge_cells(f'{celula_inicial.coordinate}:{celula_final.coordinate}')

    else:
        for coluna, valor in enumerate(args, start=1):
            sheet.cell(row=proxima_linha, column=coluna, value=valor)

    workbook.save(caminho)


def gravar_registros(tipo, registros):
    caminho = caminhos_saida.get(tipo)

    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook.active

    total = 0

    for registro in registros:
        proxima_linha = len(sheet['A']) + 1

        for coluna, valor in enumerate(registro.montar_estrutura_csv(), start=1):
            sheet.cell(row=proxima_linha, column=coluna, value=valor)

        total += registro.valor

    proxima_linha = len(sheet['A']) + 1
    celula_inicial = sheet.cell(row=proxima_linha, column=1, value='Total')
    celula_final = sheet.cell(row=proxima_linha, column=3)
    sheet.cell(row=proxima_linha, column=4, value=total)

    sheet.merge_cells(f'{celula_inicial.coordinate}:{celula_final.coordinate}')
    workbook.save(caminho)


def formatar_planilha(tipo):
    caminho = caminhos_saida.get(tipo)

    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook.active

    largura_desejada = 35
    sheet.column_dimensions['A'].width = largura_desejada
    sheet.column_dimensions['B'].width = largura_desejada
    sheet.column_dimensions['C'].width = largura_desejada
    sheet.column_dimensions['D'].width = largura_desejada
    sheet.column_dimensions['E'].width = largura_desejada

    alinhamento = Alignment(horizontal='center', vertical='center')

    for linha in sheet.iter_rows():
        for celula in linha:
            celula.alignment = alinhamento

    workbook.save(caminho)
